from __future__ import annotations

import calendar
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook, load_workbook


SKILL_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE_DIR = SKILL_ROOT / "templates"
TEMPLATE_WORKBOOK = TEMPLATE_DIR / "SAP Import JE bằng WB.xlsx"
HEADER_TXT_NAME = "Header.txt"
LINE_TXT_NAME = "Line.txt"
OUTPUT_SUBFOLDER_PREFIX = "output"

HEADER_ROW2 = [
    "JdtNum",
    "U_S1No",
    "ReferenceDate",
    "Memo",
    "Reference",
    "Reference2",
    "ProjectCode",
    "TaxDate",
    "U_VoucherTypeID",
    "U_Branch",
]

HEADER_ROW3 = [
    "JDT_NUM",
    "U_S1No",
    "RefDate",
    "Memo",
    "Ref1",
    "Ref2",
    "Project",
    "TaxDate",
    "U_VoucherTypeID",
    "U_NoteForImport",
]

LINE_ROW2 = [
    "ParentKey",
    "LineNum",
    "AccountCode",
    "Debit",
    "Credit",
    "FCDebit",
    "FCCredit",
    "FCCurrency",
    "DueDate",
    "ControlAccount",
    "ShortName",
    "LineMemo",
    "ReferenceDate1",
    "TaxDate",
    "Reference1",
    "ProjectCode",
    "CostingCode",
    "CostingCode2",
    "CostingCode3",
    "CostingCode4",
    "CostingCode5",
    "U_BankAccount",
    "BPLID",
    "U_RemarksJE",
    "BaseSum",
    "TaxGroup",
    "U_InvNo",
    "U_Invdate",
    "U_InvSeri",
    "U_InvTemplate",
    "U_isVat",
    "U_BPcode",
    "U_BPname",
    "U_TaxCode",
]

LINE_ROW3 = [
    "JdtNum",
    "LineNum",
    "Account",
    "Debit",
    "Credit",
    "FCDebit",
    "FCCredit",
    "FCCurrency",
    "DueDate",
    "Account",
    "ShortName",
    "LineMemo",
    "RefDate",
    "TaxDate",
    "Ref1",
    "Project",
    "CostingCode",
    "CostingCode2",
    "CostingCode3",
    "CostingCode4",
    "CostingCode5",
    "U_BankAccount",
    "BPLId",
    "U_RemarksJE",
    "BaseSum",
    "TaxGroup",
    "U_InvNo",
    "U_Invdate",
    "U_InvSeri",
    "U_InvTemplate",
    "U_isVat",
    "U_BPcode",
    "U_BPname",
    "U_TaxCode",
]

EXPENSE_ACCOUNT = "62721001"
VAT_ACCOUNT = "13311001"
AP_ACCOUNT = "33111001"
DISTR_RULE = "12090310;M999994;M02;PMO;M0100000"
PROJECT_CODE = "M02"
TAX_GROUP = "PVN5"
BRANCH_NOTE = 7
VOUCHER_TYPE_ID = 7012
PARTNER_CODE = "V00000162"
PARTNER_NAME = "CHI NHÁNH TỔNG CÔNG TY ĐIỆN LỰC TPHCM TNHH-CÔNG TY ĐIỆN LỰC SÀI GÒN"
PARTNER_TAX_ID = "0300951119-001"
VAT_RATE = 0.08
HEADER_SCAN_LIMIT = 20

COLUMN_ALIASES = {
    "row_no": ["STT"],
    "gross_amount": ["TONG_NO", "TỔNG NỢ", "TỔNG TIỀN", "GROSS AMOUNT"],
    "invoice_no": [
        "SỐ HD",
        "SỐ HĐ",
        "SO HD",
        "SO HĐ",
        "SERY HĐ",
        "SERY HD",
        "SỐ HÓA ĐƠN",
        "SO HOA DON",
        "INVOICE NO",
    ],
    "invoice_series": [
        "KÝ HIỆU",
        "KY HIEU",
        "SERI HĐ",
        "SERI HD",
        "MÃ KÍ HIỆU",
        "MÃ KÝ HIỆU",
        "MA KI HIEU",
        "MA KY HIEU",
        "INVOICE SERIES",
    ],
    "issue_date": [
        "NGÀY PHÁT HÀNH",
        "NGÀY PH",
        "NGAY PHAT HANH",
        "NGAY PH",
        "DOCUMENT DATE",
        "ISSUE DATE",
    ],
    "customer_code": ["MA_KHANG", "MÃ KHÁCH HÀNG", "CUSTOMER CODE"],
    "customer_name": ["TEN_KHANG", "TÊN KHÁCH HÀNG", "CUSTOMER NAME"],
    "address": ["DC_DDO", "ĐỊA CHỈ", "ADDRESS"],
    "meter_book": ["MA_SOGCS", "MÃ SỔ GCS", "METER BOOK"],
    "route_code": ["LOTRINH", "LỘ TRÌNH", "ROUTE CODE"],
    "meter_no": ["SO_DIENKE", "SỐ ĐIỆN KẾ", "SO DIEN KE", "METER NO"],
    "invoice_flag": ["HD", "HĐ", "HOA DON FLAG", "INVOICE FLAG"],
}

REQUIRED_LOGICAL_COLUMNS = ["row_no", "gross_amount", "invoice_no", "invoice_series", "issue_date"]

PROFILE_SIGNATURES = {
    "legacy_layout": {"invoice_no": {"SỐ HD", "SỐ HĐ", "SO HD", "SO HĐ"}, "invoice_series": {"KÝ HIỆU", "KY HIEU"}},
    "evn_layout_202603": {
        "invoice_no": {"SERY HĐ", "SERY HD"},
        "invoice_series": {"MÃ KÍ HIỆU", "MÃ KÝ HIỆU", "MA KI HIEU", "MA KY HIEU"},
    },
}


@dataclass
class InputRow:
    row_no: int
    customer_code: str
    customer_name: str
    address: str
    meter_book: str
    route_code: str
    meter_no: str
    invoice_flag: str
    gross_amount: int
    invoice_no: str
    invoice_series: str
    issue_date: datetime


@dataclass
class HeaderEntry:
    jdt_num: int
    reference_date: str
    memo: str
    project_code: str
    tax_date: str
    voucher_type_id: int
    branch_note: int


@dataclass
class LineEntry:
    parent_key: int
    line_num: int
    account_code: str
    debit: Optional[int]
    credit: Optional[int]
    due_date: str
    control_account: str
    short_name: str
    line_memo: str
    reference_date: str
    tax_date: str
    project_code: str
    costing_codes: list[str]
    remarks_je: str
    base_sum: Optional[int]
    tax_group: str
    invoice_no: str
    invoice_date: str
    invoice_series: str
    invoice_template: str
    is_vat: str
    bp_code: str
    bp_name: str
    tax_code: str


def normalize_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.upper().replace("Đ", "D")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def parse_datetime(value) -> datetime:
    if isinstance(value, datetime):
        return value
    if value is None or str(value).strip() == "":
        raise ValueError("Missing issue date")
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass
    raise ValueError(f"Unsupported date format: {value}")


def format_yyyymmdd(value: datetime) -> str:
    return value.strftime("%Y%m%d")


def format_ddmmyyyy(value: datetime) -> str:
    return value.strftime("%d/%m/%Y")


def previous_month_period(issue_date: datetime) -> tuple[int, int]:
    year = issue_date.year
    month = issue_date.month - 1
    if month == 0:
        month = 12
        year -= 1
    return month, year


def build_header_memo(issue_date: datetime) -> str:
    month, year = previous_month_period(issue_date)
    return f"Điện tiêu thụ tháng {month} năm {year}"



def build_description(issue_date: datetime) -> str:
    month, year = previous_month_period(issue_date)
    last_day = calendar.monthrange(year, month)[1]
    return (
        f"Điện tiêu thụ tháng {month} năm {year} "
        f"từ ngày 01/{month:02d}/{year} đến ngày {last_day:02d}/{month:02d}/{year}"
    )


def truncate_text(value: str, max_length: int) -> str:
    text = normalize_text(value)
    if len(text) <= max_length:
        return text
    return text[:max_length]


def split_costing_codes(distr_rule: str) -> list[str]:
    parts = [part.strip() for part in distr_rule.split(";")]
    while len(parts) < 5:
        parts.append("")
    return parts[:5]


def infer_period_label(rows: List[InputRow]) -> tuple[str, str]:
    if not rows:
        raise ValueError("Cannot infer period label from empty rows")
    month, year = previous_month_period(rows[0].issue_date)
    return f"{month:02d}", f"{year}"


def infer_output_stem(rows: List[InputRow]) -> str:
    month_label, year_label = infer_period_label(rows)
    return f"SAP_Import by JE_{month_label}_{year_label}"


def resolve_input_file(path_arg: str) -> Path:
    path = Path(path_arg)
    if path.is_file():
        return path
    if not path.exists():
        raise FileNotFoundError(f"Input path not found: {path}")
    candidates = [
        p
        for p in path.iterdir()
        if p.is_file() and p.suffix.lower() == ".xlsx" and "sap" not in p.name.lower() and not p.name.startswith("~$")
    ]
    if not candidates:
        raise FileNotFoundError("No input .xlsx file found in folder")
    candidates.sort(key=lambda p: p.name.lower())
    return candidates[0]


def resolve_output_file(input_file: Path, rows: List[InputRow], explicit_output: Optional[str]) -> Path:
    if explicit_output:
        output_file = Path(explicit_output)
    else:
        month_label, year_label = infer_period_label(rows)
        output_dir = input_file.parent / f"{OUTPUT_SUBFOLDER_PREFIX}_{month_label}_{year_label}"
        output_file = output_dir / f"{infer_output_stem(rows)}.xlsx"
    output_file.parent.mkdir(parents=True, exist_ok=True)
    return output_file


def build_alias_lookup() -> dict[str, str]:
    lookup: dict[str, str] = {}
    for logical_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            lookup[normalize_header(alias)] = logical_name
    return lookup


def detect_header_row(ws) -> tuple[int, dict[str, int], dict[str, str]]:
    alias_lookup = build_alias_lookup()
    best_row = None
    best_mapping: dict[str, int] = {}
    best_headers: dict[str, str] = {}

    max_scan_row = min(ws.max_row, HEADER_SCAN_LIMIT)
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan_row, values_only=True), start=1):
        logical_to_index: dict[str, int] = {}
        logical_to_header: dict[str, str] = {}
        for col_idx, cell_value in enumerate(row):
            normalized = normalize_header(cell_value)
            if not normalized:
                continue
            logical_name = alias_lookup.get(normalized)
            if logical_name and logical_name not in logical_to_index:
                logical_to_index[logical_name] = col_idx
                logical_to_header[logical_name] = normalize_text(cell_value)

        score = len(logical_to_index)
        required_hits = sum(1 for name in REQUIRED_LOGICAL_COLUMNS if name in logical_to_index)
        if required_hits >= 4 and score >= 5:
            return row_idx, logical_to_index, logical_to_header
        if best_row is None or required_hits > sum(1 for name in REQUIRED_LOGICAL_COLUMNS if name in best_mapping) or (
            required_hits == sum(1 for name in REQUIRED_LOGICAL_COLUMNS if name in best_mapping) and score > len(best_mapping)
        ):
            best_row = row_idx
            best_mapping = logical_to_index
            best_headers = logical_to_header

    missing = [name for name in REQUIRED_LOGICAL_COLUMNS if name not in best_mapping]
    raise ValueError(
        "Could not detect a valid header row. "
        f"Best candidate row: {best_row}, missing required columns: {missing}"
    )


def detect_profile(header_names: dict[str, str]) -> str:
    normalized_headers = {name: normalize_header(value) for name, value in header_names.items()}
    for profile_name, signature in PROFILE_SIGNATURES.items():
        matched = True
        for logical_name, accepted in signature.items():
            if normalized_headers.get(logical_name, "") not in accepted:
                matched = False
                break
        if matched:
            return profile_name
    return "generic_header_mapping"


def get_cell(row_values: tuple, column_map: dict[str, int], logical_name: str):
    idx = column_map.get(logical_name)
    if idx is None or idx >= len(row_values):
        return None
    return row_values[idx]


def parse_row_no(value) -> Optional[int]:
    if value in (None, ""):
        return None
    text = normalize_text(value)
    if not text:
        return None
    if normalize_header(text) == "TONG CONG":
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def parse_gross_amount(value, row_no: int) -> int:
    if value in (None, ""):
        raise ValueError(f"Missing TONG_NO at STT {row_no}")
    text = normalize_text(value).replace(",", "")
    return int(round(float(text)))


def validate_required_text(value, field_label: str, row_no: int) -> str:
    text = normalize_text(value)
    if not text:
        raise ValueError(f"Missing {field_label} at STT {row_no}")
    return text


def read_input_rows(input_file: Path) -> tuple[str, List[InputRow], dict]:
    wb = load_workbook(input_file, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title
    header_row, column_map, detected_headers = detect_header_row(ws)
    profile = detect_profile(detected_headers)

    rows: List[InputRow] = []
    skipped_rows = 0
    for excel_row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        row_no = parse_row_no(get_cell(excel_row, column_map, "row_no"))
        if row_no is None:
            if any(normalize_text(v) for v in excel_row):
                skipped_rows += 1
            continue

        invoice_no = validate_required_text(get_cell(excel_row, column_map, "invoice_no"), "invoice_no", row_no)
        invoice_series = validate_required_text(get_cell(excel_row, column_map, "invoice_series"), "invoice_series", row_no)
        issue_date = parse_datetime(get_cell(excel_row, column_map, "issue_date"))
        gross_amount = parse_gross_amount(get_cell(excel_row, column_map, "gross_amount"), row_no)

        rows.append(
            InputRow(
                row_no=row_no,
                customer_code=normalize_text(get_cell(excel_row, column_map, "customer_code")),
                customer_name=normalize_text(get_cell(excel_row, column_map, "customer_name")),
                address=normalize_text(get_cell(excel_row, column_map, "address")),
                meter_book=normalize_text(get_cell(excel_row, column_map, "meter_book")),
                route_code=normalize_text(get_cell(excel_row, column_map, "route_code")),
                meter_no=normalize_text(get_cell(excel_row, column_map, "meter_no")),
                invoice_flag=normalize_text(get_cell(excel_row, column_map, "invoice_flag")),
                gross_amount=gross_amount,
                invoice_no=invoice_no,
                invoice_series=invoice_series,
                issue_date=issue_date,
            )
        )
    if not rows:
        raise ValueError("No usable invoice rows found")
    metadata = {
        "sheet_name": sheet_name,
        "header_row": header_row,
        "input_profile": profile,
        "detected_headers": detected_headers,
        "skipped_rows": skipped_rows,
    }
    return sheet_name, rows, metadata


def build_header_entry(rows: List[InputRow], jdt_num: int, import_created_at: datetime) -> HeaderEntry:
    if not rows:
        raise ValueError("Cannot build header entry from empty rows")
    header_source_row = max(rows, key=lambda item: item.issue_date)
    posting_date = format_yyyymmdd(import_created_at)
    header_memo = truncate_text(build_header_memo(header_source_row.issue_date), 50)
    return HeaderEntry(
        jdt_num=jdt_num,
        reference_date=posting_date,
        memo=header_memo,
        project_code=PROJECT_CODE,
        tax_date=posting_date,
        voucher_type_id=VOUCHER_TYPE_ID,
        branch_note=BRANCH_NOTE,
    )



def build_entries(
    item: InputRow,
    jdt_num: int,
    line_start: int,
    header_memo: str,
    header_reference_date: str,
) -> list[LineEntry]:
    base_amount = int(round(item.gross_amount / (1 + VAT_RATE)))
    vat_amount = item.gross_amount - base_amount
    description = build_description(item.issue_date)
    posting_date = header_reference_date
    line_tax_date = format_yyyymmdd(item.issue_date)
    invoice_date = format_ddmmyyyy(item.issue_date)
    costing_codes = split_costing_codes(DISTR_RULE)

    expense = LineEntry(
        parent_key=jdt_num,
        line_num=line_start,
        account_code=EXPENSE_ACCOUNT,
        debit=base_amount,
        credit=None,
        due_date=posting_date,
        control_account=EXPENSE_ACCOUNT,
        short_name="",
        line_memo=header_memo,
        reference_date=posting_date,
        tax_date=line_tax_date,
        project_code=PROJECT_CODE,
        costing_codes=costing_codes,
        remarks_je=description,
        base_sum=None,
        tax_group="",
        invoice_no="",
        invoice_date="",
        invoice_series="",
        invoice_template="",
        is_vat="",
        bp_code="",
        bp_name="",
        tax_code="",
    )
    vat = LineEntry(
        parent_key=jdt_num,
        line_num=line_start + 1,
        account_code=VAT_ACCOUNT,
        debit=vat_amount,
        credit=None,
        due_date=posting_date,
        control_account=VAT_ACCOUNT,
        short_name="",
        line_memo=header_memo,
        reference_date=posting_date,
        tax_date=line_tax_date,
        project_code=PROJECT_CODE,
        costing_codes=["", "", "", "", ""],
        remarks_je=description,
        base_sum=item.gross_amount,
        tax_group=TAX_GROUP,
        invoice_no=item.invoice_no,
        invoice_date=invoice_date,
        invoice_series=item.invoice_series,
        invoice_template="",
        is_vat="Y",
        bp_code=PARTNER_CODE,
        bp_name=PARTNER_NAME,
        tax_code=PARTNER_TAX_ID,
    )
    ap = LineEntry(
        parent_key=jdt_num,
        line_num=line_start + 2,
        account_code=AP_ACCOUNT,
        debit=None,
        credit=item.gross_amount,
        due_date=posting_date,
        control_account=AP_ACCOUNT,
        short_name=PARTNER_CODE,
        line_memo=header_memo,
        reference_date=posting_date,
        tax_date=line_tax_date,
        project_code=PROJECT_CODE,
        costing_codes=["", "", "", "", ""],
        remarks_je=description,
        base_sum=None,
        tax_group="",
        invoice_no="",
        invoice_date="",
        invoice_series="",
        invoice_template="",
        is_vat="",
        bp_code="",
        bp_name="",
        tax_code="",
    )
    return [expense, vat, ap]


def load_template_workbook() -> Workbook:
    if not TEMPLATE_WORKBOOK.exists():
        raise FileNotFoundError(f"Template workbook not found: {TEMPLATE_WORKBOOK}")
    return load_workbook(TEMPLATE_WORKBOOK)


def reset_template_sheets(wb: Workbook) -> tuple:
    if "JE-Header" not in wb.sheetnames or "JE-Line" not in wb.sheetnames:
        raise ValueError("Template workbook must contain JE-Header and JE-Line sheets")
    ws_header = wb["JE-Header"]
    ws_line = wb["JE-Line"]
    if ws_header.max_row > 3:
        ws_header.delete_rows(4, ws_header.max_row - 3)
    if ws_line.max_row > 3:
        ws_line.delete_rows(4, ws_line.max_row - 3)

    for idx, value in enumerate(HEADER_ROW2, start=1):
        ws_header.cell(2, idx).value = value
    for idx, value in enumerate(HEADER_ROW3, start=1):
        ws_header.cell(3, idx).value = value

    for idx, value in enumerate(LINE_ROW2, start=1):
        ws_line.cell(2, idx).value = value
    for idx, value in enumerate(LINE_ROW3, start=1):
        ws_line.cell(3, idx).value = value

    return ws_header, ws_line


def write_workbook(header_rows: list[HeaderEntry], line_rows: list[LineEntry], output_file: Path) -> None:
    wb = load_template_workbook()
    ws_header, ws_line = reset_template_sheets(wb)

    for offset, item in enumerate(header_rows, start=4):
        ws_header.cell(offset, 1).value = item.jdt_num
        ws_header.cell(offset, 2).value = ""
        ws_header.cell(offset, 3).value = item.reference_date
        ws_header.cell(offset, 4).value = item.memo
        ws_header.cell(offset, 5).value = ""
        ws_header.cell(offset, 6).value = ""
        ws_header.cell(offset, 7).value = item.project_code
        ws_header.cell(offset, 8).value = item.tax_date
        ws_header.cell(offset, 9).value = item.voucher_type_id
        ws_header.cell(offset, 10).value = item.branch_note

    for offset, item in enumerate(line_rows, start=4):
        ws_line.cell(offset, 1).value = item.parent_key
        ws_line.cell(offset, 2).value = item.line_num
        ws_line.cell(offset, 3).value = item.account_code
        ws_line.cell(offset, 4).value = item.debit
        ws_line.cell(offset, 5).value = item.credit
        ws_line.cell(offset, 6).value = 0
        ws_line.cell(offset, 7).value = 0
        ws_line.cell(offset, 8).value = ""
        ws_line.cell(offset, 9).value = item.due_date
        ws_line.cell(offset, 10).value = item.control_account
        ws_line.cell(offset, 11).value = item.short_name
        ws_line.cell(offset, 12).value = item.line_memo
        ws_line.cell(offset, 13).value = item.reference_date
        ws_line.cell(offset, 14).value = item.tax_date
        ws_line.cell(offset, 15).value = ""
        ws_line.cell(offset, 16).value = item.project_code
        ws_line.cell(offset, 17).value = item.costing_codes[0]
        ws_line.cell(offset, 18).value = item.costing_codes[1]
        ws_line.cell(offset, 19).value = item.costing_codes[2]
        ws_line.cell(offset, 20).value = item.costing_codes[3]
        ws_line.cell(offset, 21).value = item.costing_codes[4]
        ws_line.cell(offset, 22).value = ""
        ws_line.cell(offset, 23).value = BRANCH_NOTE
        ws_line.cell(offset, 24).value = item.remarks_je
        ws_line.cell(offset, 25).value = item.base_sum
        ws_line.cell(offset, 26).value = item.tax_group
        ws_line.cell(offset, 27).value = item.invoice_no
        ws_line.cell(offset, 28).value = item.invoice_date
        ws_line.cell(offset, 29).value = item.invoice_series
        ws_line.cell(offset, 30).value = item.invoice_template
        ws_line.cell(offset, 31).value = item.is_vat
        ws_line.cell(offset, 32).value = item.bp_code
        ws_line.cell(offset, 33).value = item.bp_name
        ws_line.cell(offset, 34).value = item.tax_code

    wb.save(output_file)


def rows_to_tsv(rows: list[list[object]]) -> str:
    def serialize(value: object) -> str:
        if value is None:
            return ""
        return str(value)
    return "\n".join("\t".join(serialize(value) for value in row) for row in rows) + "\n"


def write_text_exports(header_rows: list[HeaderEntry], line_rows: list[LineEntry], output_file: Path) -> tuple[Path, Path]:
    output_dir = output_file.parent
    header_txt = output_dir / HEADER_TXT_NAME
    line_txt = output_dir / LINE_TXT_NAME

    header_tsv_rows: list[list[object]] = [HEADER_ROW2, HEADER_ROW3]
    for item in header_rows:
        header_tsv_rows.append([
            item.jdt_num,
            "",
            item.reference_date,
            item.memo,
            "",
            "",
            item.project_code,
            item.tax_date,
            item.voucher_type_id,
            item.branch_note,
        ])

    line_tsv_rows: list[list[object]] = [LINE_ROW2, LINE_ROW3]
    for item in line_rows:
        line_tsv_rows.append([
            item.parent_key,
            item.line_num,
            item.account_code,
            item.debit,
            item.credit,
            0,
            0,
            "",
            item.due_date,
            item.control_account,
            item.short_name,
            item.line_memo,
            item.reference_date,
            item.tax_date,
            "",
            item.project_code,
            item.costing_codes[0],
            item.costing_codes[1],
            item.costing_codes[2],
            item.costing_codes[3],
            item.costing_codes[4],
            "",
            BRANCH_NOTE,
            item.remarks_je,
            item.base_sum,
            item.tax_group,
            item.invoice_no,
            item.invoice_date,
            item.invoice_series,
            item.invoice_template,
            item.is_vat,
            item.bp_code,
            item.bp_name,
            item.tax_code,
        ])

    header_txt.write_text(rows_to_tsv(header_tsv_rows), encoding="utf-16")
    line_txt.write_text(rows_to_tsv(line_tsv_rows), encoding="utf-16")
    return header_txt, line_txt


def validate(rows: List[InputRow], header_rows: List[HeaderEntry], line_rows: List[LineEntry]) -> dict:
    input_total = sum(r.gross_amount for r in rows)
    debit_total = sum(item.debit or 0 for item in line_rows)
    credit_total = sum(item.credit or 0 for item in line_rows)

    if input_total != debit_total:
        raise ValueError(f"Control failed: input total {input_total} != output debit total {debit_total}")
    if input_total != credit_total:
        raise ValueError(f"Control failed: input total {input_total} != output credit total {credit_total}")
    if len(header_rows) != 1:
        raise ValueError(f"Control failed: expected exactly 1 header row, got {len(header_rows)}")
    if len(line_rows) != len(rows) * 3:
        raise ValueError("Control failed: output row count is not 3x input row count")

    expected_line_nums = list(range(len(line_rows)))
    actual_line_nums = [item.line_num for item in line_rows]
    if actual_line_nums != expected_line_nums:
        raise ValueError("Control failed: line numbers are not sequential for single journal output")

    parent_keys = {item.parent_key for item in line_rows}
    if parent_keys != {header_rows[0].jdt_num}:
        raise ValueError("Control failed: line parent keys do not match the single header journal number")

    missing_invoice_no = sum(1 for r in rows if not r.invoice_no)
    missing_invoice_series = sum(1 for r in rows if not r.invoice_series)
    if missing_invoice_no or missing_invoice_series:
        raise ValueError(
            "Control failed: missing invoice fields in input. "
            f"missing invoice_no={missing_invoice_no}, missing invoice_series={missing_invoice_series}"
        )

    return {
        "invoice_count": len(rows),
        "header_row_count": len(header_rows),
        "line_row_count": len(line_rows),
        "output_row_count": len(line_rows),
        "input_total": input_total,
        "debit_total": debit_total,
        "credit_total": credit_total,
        "template_format": "SAP Import JE bằng WB",
        "output_mode": "single_header_multi_invoice_lines",
    }


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print("Usage: python build_sap_electricity_import.py <input-file-or-folder> [output-file]", file=sys.stderr)
        return 1

    input_file = resolve_input_file(argv[1])
    sheet_name, rows, metadata = read_input_rows(input_file)
    output_file = resolve_output_file(input_file, rows, argv[2] if len(argv) >= 3 else None)

    import_created_at = datetime.now()
    journal_number = 1
    header_rows: list[HeaderEntry] = [build_header_entry(rows, journal_number, import_created_at)]
    header_memo = header_rows[0].memo
    header_reference_date = header_rows[0].reference_date
    line_rows: list[LineEntry] = []
    for index, item in enumerate(rows):
        line_start = index * 3
        journal_lines = build_entries(item, journal_number, line_start, header_memo, header_reference_date)
        line_rows.extend(journal_lines)

    summary = validate(rows, header_rows, line_rows)

    write_workbook(header_rows, line_rows, output_file)
    header_txt, line_txt = write_text_exports(header_rows, line_rows, output_file)

    month_label, year_label = infer_period_label(rows)

    summary.update(metadata)
    summary["period_month"] = month_label
    summary["period_year"] = year_label
    summary["input_file"] = str(input_file)
    summary["output_file"] = str(output_file)
    summary["output_folder"] = str(output_file.parent)
    summary["header_txt_file"] = str(header_txt)
    summary["line_txt_file"] = str(line_txt)
    summary["sheet_name"] = sheet_name
    summary["template_workbook"] = str(TEMPLATE_WORKBOOK)
    summary["ap_account"] = AP_ACCOUNT
    summary["header_date_rule"] = "import_created_at"
    summary["import_created_at"] = import_created_at.strftime("%Y-%m-%d %H:%M:%S")

    summary_json = json.dumps(summary, ensure_ascii=False, indent=2)
    summary_file = output_file.with_suffix(".summary.json")
    summary_file.write_text(summary_json, encoding="utf-8")

    try:
        print(summary_json)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((summary_json + "\n").encode("utf-8", errors="replace"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv))
