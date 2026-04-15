#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import json
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.styles import PatternFill

SKILL_DIR = Path(__file__).resolve().parents[1]

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass

EXPENSE_ACCOUNT = '64281001'
VAT_ACCOUNT = '13311001'
AP_ACCOUNT = '33111001'
VENDOR_CODE = 'V00000070'
VENDOR_NAME = 'CÔNG TY TNHH GRAB'
VENDOR_TAX_CODE = '312650437'
VENDOR_ADDRESS = '268 Tô Hiến Thành, Thành phố Hồ Chí Minh, Quận 10'
TAX_GROUP = 'PVN5'
DEFAULT_COSTING_STRING = '17020101;M999998;M02;ADM;M0100000'
GRAB_PREFIX = 'GFB Billing Calculation Report'
HIGHLIGHT_FILL = PatternFill(fill_type='solid', fgColor='FFF59D')
TEMPLATE_CANDIDATES = [
    SKILL_DIR / 'SAP_Template import JE bằng WB.xlsx',
]
REQUIRED_COLUMNS = [
    'COMPANY_NAME', 'GROUP_NAME', 'AMOUNT', 'PRE_VAT_DELIVERY_FEE', 'VAT_VALUE_DELIVERY_FEE',
    'PRE_VAT_SERVICE_FEE', 'VAT_VALUE_SERVICE_FEE', 'INVOICE_NUMBER', 'VAT_INVOICE_DATE',
    'VAT_INVOICE_SERIAL', 'TRANSACTION_TIME'
]
HEADER_COLUMNS = ['JdtNum', 'U_S1No', 'ReferenceDate', 'Memo', 'Reference', 'Reference2', 'ProjectCode', 'TaxDate', 'U_VoucherTypeID', 'U_Branch']
LINE_COLUMNS = [
    'ParentKey', 'LineNum', 'AccountCode', 'Debit', 'Credit', 'FCDebit', 'FCCredit', 'FCCurrency',
    'DueDate', 'ControlAccount', 'ShortName', 'LineMemo', 'ReferenceDate1', 'TaxDate', 'Reference1', 'ProjectCode',
    'CostingCode', 'CostingCode2', 'CostingCode3', 'CostingCode4', 'CostingCode5', 'U_BankAccount',
    'BPLID', 'U_RemarksJE', 'BaseSum', 'TaxGroup', 'U_InvNo', 'U_Invdate', 'U_InvSeri', 'U_InvTemplate',
    'U_isVat', 'U_BPcode', 'U_BPname', 'U_TaxCode'
]


@dataclass
class TemplateDefaults:
    header_project: Any
    header_voucher_type: Any
    header_branch: Any
    line_project: Any
    line_bplid: Any
    line_costing2: Any
    line_costing3: Any
    line_costing4: Any
    line_costing5: Any


@dataclass
class ConversionResult:
    output_path: str
    summary_path: str
    header_txt_path: str
    line_txt_path: str
    template_path: str
    billing_month: str
    posting_date: str
    gfb_rows: int
    header_rows: int
    line_rows: int
    department_count: int
    total_amount: int
    total_debit: int
    total_credit: int
    warnings: list[str]
    department_summary: list[dict[str, Any]]


@dataclass
class BuildArtifacts:
    header_rows: list[dict[str, Any]]
    line_rows: list[dict[str, Any]]
    warnings: list[str]
    total_debit: int
    total_credit: int
    processed_rows: int
    department_summary: list[dict[str, Any]]
    highlight_header_keys: set[int]
    highlight_line_keys: set[tuple[int, int]]


def to_num(value: Any) -> float:
    try:
        return float(value) if value not in (None, '') else 0.0
    except (TypeError, ValueError):
        return 0.0



def clean_text(value: Any) -> str:
    return '' if value is None else str(value).strip()



def format_amount(value: float) -> Optional[int]:
    rounded = round(value)
    return rounded if rounded != 0 else None



def yyyymmdd_int(d: date) -> int:
    return int(d.strftime('%Y%m%d'))



def parse_date(value: Any) -> Optional[date]:
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(text[:10], fmt).date()
        except ValueError:
            pass
    return None



def strip_serial_prefix(serial: Any) -> Optional[str]:
    text = clean_text(serial)
    if not text:
        return None
    return text[1:] if text.startswith('1') else text



def get_last_day_of_next_month(year: int, month: int) -> date:
    if month == 12:
        year, month = year + 1, 1
    else:
        month += 1
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)



def detect_billing_month(rows: list[dict[str, Any]]) -> tuple[int, int]:
    months = []
    for row in rows:
        for candidate in (row.get('VAT_INVOICE_DATE'), row.get('TRANSACTION_TIME')):
            d = parse_date(candidate)
            if d:
                months.append((d.year, d.month))
                break
    if not months:
        today = date.today()
        return today.year, today.month
    return Counter(months).most_common(1)[0][0]



def split_costing(costing_text: str) -> list[Optional[str]]:
    parts = [clean_text(x) or None for x in costing_text.split(';')]
    while len(parts) < 5:
        parts.append(None)
    return parts[:5]



def truncate(text: str, max_len: int) -> str:
    text = clean_text(text)
    return text if len(text) <= max_len else text[: max_len - 3].rstrip() + '...'



def build_header_memo(month_str: str, department_code: str, invoice_count: int) -> str:
    base = f'Chi phí Grab T{month_str} {department_code}'
    if invoice_count > 1:
        base += f' ({invoice_count} HĐ)'
    return truncate(base, 50)



def build_line_memo(month_str: str, invoice_no: Any) -> str:
    invoice = clean_text(invoice_no)
    base = f'Chi phí Grab T{month_str}'
    if invoice:
        base += f' HĐ {invoice}'
    return truncate(base, 50)



def resolve_input_path(input_path: Path) -> Path:
    if input_path.is_file():
        return input_path
    candidates = sorted([p for p in input_path.glob('*.xlsx') if p.name.startswith(GRAB_PREFIX)])
    if not candidates:
        raise FileNotFoundError(f'Không tìm thấy file GFB Billing Calculation Report trong {input_path}')
    return candidates[0]



def resolve_template_path(template: Optional[str]) -> Path:
    if template:
        path = Path(template)
        if not path.exists():
            raise FileNotFoundError(f'Không tìm thấy template: {path}')
        return path
    for candidate in TEMPLATE_CANDIDATES:
        if candidate.exists():
            return candidate
    raise FileNotFoundError('Không tìm thấy template SAP_Template import JE bằng WB.xlsx')



def resolve_output_path(base_folder: Path, billing_month: str, output: Optional[str]) -> Path:
    if output:
        output_candidate = Path(output)
        if output_candidate.suffix.lower() == '.xlsx':
            output_dir = output_candidate.parent / 'Output'
            output_name = output_candidate.name
        else:
            output_dir = output_candidate / 'Output'
            output_name = f'SAP JE Import_Grab_{billing_month}.xlsx'
    else:
        output_dir = base_folder / 'Output'
        output_name = f'SAP JE Import_Grab_{billing_month}.xlsx'
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / output_name



def read_gfb_rows(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in row):
            continue
        row_dict = dict(zip(headers, row))
        if to_num(row_dict.get('AMOUNT')) > 0 and clean_text(row_dict.get('COMPANY_NAME')):
            rows.append(row_dict)
    missing = [col for col in REQUIRED_COLUMNS if col not in headers]
    if missing:
        raise ValueError(f'Thiếu cột bắt buộc trong file GFB: {missing}')
    return rows



def load_template_defaults(path: Path) -> TemplateDefaults:
    wb = openpyxl.load_workbook(path)
    header = wb['JE-Header']
    line = wb['JE-Line']
    defaults = TemplateDefaults(
        header_project=header.cell(4, 7).value,
        header_voucher_type=header.cell(4, 9).value,
        header_branch=header.cell(4, 10).value,
        line_project=line.cell(4, 15).value,
        line_bplid=line.cell(4, 22).value,
        line_costing2=line.cell(4, 17).value,
        line_costing3=line.cell(4, 18).value,
        line_costing4=line.cell(4, 19).value,
        line_costing5=line.cell(4, 20).value,
    )
    wb.close()
    return defaults



def clear_sheet_data(ws, from_row: int, max_col: int) -> None:
    for row in range(from_row, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.fill = PatternFill(fill_type=None)



def ensure_line_taxdate_column(ws) -> None:
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    if 'TaxDate' in headers:
        return

    reference_date_idx = headers.index('ReferenceDate1') + 1
    insert_at = reference_date_idx + 1
    ws.insert_cols(insert_at)

    ws.cell(1, insert_at).value = 'Nhập ngày hóa đơn, format: YYYYMMDD'
    ws.cell(2, insert_at).value = 'TaxDate'
    ws.cell(3, insert_at).value = 'TaxDate'



def normalize_department_code(group_name: Any) -> tuple[str, bool]:
    code = clean_text(group_name)
    if code:
        return code, False
    return 'General', True



def build_rows(gfb_rows: list[dict[str, Any]], posting_date: date, month_str: str, defaults: TemplateDefaults) -> BuildArtifacts:
    grouped_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    blank_group_count = 0

    for row in gfb_rows:
        department_code, was_blank = normalize_department_code(row.get('GROUP_NAME'))
        row_copy = dict(row)
        row_copy['_department_code'] = department_code
        row_copy['_group_was_blank'] = was_blank
        grouped_rows[department_code].append(row_copy)
        if was_blank:
            blank_group_count += 1

    header_rows: list[dict[str, Any]] = []
    line_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    total_debit = 0
    total_credit = 0
    posting_int = yyyymmdd_int(posting_date)
    costing = split_costing(DEFAULT_COSTING_STRING)
    highlight_header_keys: set[int] = set()
    highlight_line_keys: set[tuple[int, int]] = set()
    department_summary: list[dict[str, Any]] = []

    if blank_group_count:
        warnings.append(f'{blank_group_count} dòng GROUP_NAME trống được map vào General và tô vàng trên Excel')

    sorted_departments = sorted(grouped_rows.items(), key=lambda item: item[0])

    for jdt_num, (department_code, department_rows) in enumerate(sorted_departments, start=1):
        sorted_rows = sorted(
            department_rows,
            key=lambda r: (
                clean_text(r.get('VAT_INVOICE_DATE')),
                clean_text(r.get('INVOICE_NUMBER')),
                clean_text(r.get('TRANSACTION_TIME')),
            ),
        )
        header_memo = build_header_memo(month_str, department_code, len(sorted_rows))
        header_has_highlight = any(bool(r.get('_group_was_blank')) for r in sorted_rows)

        header_rows.append({
            'JdtNum': jdt_num,
            'U_S1No': None,
            'ReferenceDate': posting_int,
            'Memo': header_memo,
            'Reference': None,
            'Reference2': None,
            'ProjectCode': defaults.header_project,
            'TaxDate': posting_int,
            'U_VoucherTypeID': defaults.header_voucher_type,
            'U_Branch': defaults.header_branch,
        })
        if header_has_highlight:
            highlight_header_keys.add(jdt_num)

        department_amount = 0
        department_pre_vat = 0
        department_vat = 0
        current_line_num = 0

        for row in sorted_rows:
            invoice_no = clean_text(row.get('INVOICE_NUMBER')) or None
            invoice_date = parse_date(row.get('VAT_INVOICE_DATE'))
            invoice_tax_date = yyyymmdd_int(invoice_date) if invoice_date else None
            invoice_serial = strip_serial_prefix(row.get('VAT_INVOICE_SERIAL'))
            pre_vat = round(to_num(row.get('PRE_VAT_DELIVERY_FEE')) + to_num(row.get('PRE_VAT_SERVICE_FEE')))
            vat_amt = round(to_num(row.get('VAT_VALUE_DELIVERY_FEE')) + to_num(row.get('VAT_VALUE_SERVICE_FEE')))
            total = pre_vat + vat_amt
            line_memo = build_line_memo(month_str, invoice_no)
            row_highlight = bool(row.get('_group_was_blank'))

            if not invoice_no:
                warnings.append(f'JE {jdt_num} / {department_code}: thiếu INVOICE_NUMBER')
            if not invoice_date:
                warnings.append(f'JE {jdt_num} / {department_code}: thiếu VAT_INVOICE_DATE')
            if vat_amt == 0:
                warnings.append(f'JE {jdt_num} / {department_code}: VAT = 0 / HĐ {invoice_no or "(blank)"}')

            expense_line_num = current_line_num
            line_rows.append({
                'ParentKey': jdt_num,
                'LineNum': expense_line_num,
                'AccountCode': EXPENSE_ACCOUNT,
                'Debit': format_amount(pre_vat),
                'Credit': None,
                'FCDebit': 0,
                'FCCredit': 0,
                'FCCurrency': None,
                'DueDate': posting_int,
                'ControlAccount': EXPENSE_ACCOUNT,
                'ShortName': None,
                'LineMemo': line_memo,
                'ReferenceDate1': posting_int,
                'TaxDate': invoice_tax_date,
                'Reference1': None,
                'ProjectCode': defaults.line_project,
                'CostingCode': costing[0],
                'CostingCode2': costing[1] or defaults.line_costing2,
                'CostingCode3': costing[2] or defaults.line_costing3,
                'CostingCode4': costing[3] or defaults.line_costing4,
                'CostingCode5': costing[4] or defaults.line_costing5,
                'U_BankAccount': None,
                'BPLID': defaults.line_bplid,
                'U_RemarksJE': truncate(line_memo, 200),
                'BaseSum': None,
                'TaxGroup': None,
                'U_InvNo': None,
                'U_Invdate': None,
                'U_InvSeri': None,
                'U_InvTemplate': None,
                'U_isVat': None,
                'U_BPcode': None,
                'U_BPname': None,
                'U_TaxCode': None,
            })
            if row_highlight:
                highlight_line_keys.add((jdt_num, expense_line_num))
            current_line_num += 1
            total_debit += pre_vat
            department_pre_vat += pre_vat

            if vat_amt:
                vat_line_num = current_line_num
                line_rows.append({
                    'ParentKey': jdt_num,
                    'LineNum': vat_line_num,
                    'AccountCode': VAT_ACCOUNT,
                    'Debit': format_amount(vat_amt),
                    'Credit': None,
                    'FCDebit': 0,
                    'FCCredit': 0,
                    'FCCurrency': None,
                    'DueDate': posting_int,
                    'ControlAccount': VAT_ACCOUNT,
                    'ShortName': None,
                    'LineMemo': line_memo,
                    'ReferenceDate1': posting_int,
                    'TaxDate': invoice_tax_date,
                    'Reference1': None,
                    'ProjectCode': defaults.line_project,
                    'CostingCode': None,
                    'CostingCode2': None,
                    'CostingCode3': None,
                    'CostingCode4': None,
                    'CostingCode5': None,
                    'U_BankAccount': None,
                    'BPLID': defaults.line_bplid,
                    'U_RemarksJE': truncate(line_memo, 200),
                    'BaseSum': format_amount(pre_vat),
                    'TaxGroup': TAX_GROUP,
                    'U_InvNo': invoice_no,
                    'U_Invdate': invoice_date,
                    'U_InvSeri': invoice_serial,
                    'U_InvTemplate': None,
                    'U_isVat': 'Y',
                    'U_BPcode': VENDOR_CODE,
                    'U_BPname': VENDOR_NAME,
                    'U_TaxCode': VENDOR_TAX_CODE,
                })
                if row_highlight:
                    highlight_line_keys.add((jdt_num, vat_line_num))
                current_line_num += 1
                total_debit += vat_amt
                department_vat += vat_amt

            ap_line_num = current_line_num
            line_rows.append({
                'ParentKey': jdt_num,
                'LineNum': ap_line_num,
                'AccountCode': AP_ACCOUNT,
                'Debit': None,
                'Credit': format_amount(total),
                'FCDebit': 0,
                'FCCredit': 0,
                'FCCurrency': None,
                'DueDate': posting_int,
                'ControlAccount': AP_ACCOUNT,
                'ShortName': VENDOR_CODE,
                'LineMemo': line_memo,
                'ReferenceDate1': posting_int,
                'TaxDate': invoice_tax_date,
                'Reference1': None,
                'ProjectCode': defaults.line_project,
                'CostingCode': None,
                'CostingCode2': None,
                'CostingCode3': None,
                'CostingCode4': None,
                'CostingCode5': None,
                'U_BankAccount': None,
                'BPLID': defaults.line_bplid,
                'U_RemarksJE': truncate(line_memo, 200),
                'BaseSum': None,
                'TaxGroup': None,
                'U_InvNo': None,
                'U_Invdate': None,
                'U_InvSeri': None,
                'U_InvTemplate': None,
                'U_isVat': None,
                'U_BPcode': None,
                'U_BPname': None,
                'U_TaxCode': None,
            })
            if row_highlight:
                highlight_line_keys.add((jdt_num, ap_line_num))
            current_line_num += 1
            total_credit += total
            department_amount += round(to_num(row.get('AMOUNT')))

        department_summary.append({
            'department_code': department_code,
            'invoice_count': len(sorted_rows),
            'highlighted_invoice_count': sum(1 for r in sorted_rows if r.get('_group_was_blank')),
            'pre_vat': department_pre_vat,
            'vat': department_vat,
            'amount': department_amount,
            'jdt_num': jdt_num,
        })

    return BuildArtifacts(
        header_rows=header_rows,
        line_rows=line_rows,
        warnings=warnings,
        total_debit=total_debit,
        total_credit=total_credit,
        processed_rows=len(gfb_rows),
        department_summary=department_summary,
        highlight_header_keys=highlight_header_keys,
        highlight_line_keys=highlight_line_keys,
    )



def write_output(template_path: Path, output_path: Path, header_rows: list[dict[str, Any]], line_rows: list[dict[str, Any]], highlight_header_keys: set[int], highlight_line_keys: set[tuple[int, int]]) -> None:
    wb = openpyxl.load_workbook(template_path)
    header_ws = wb['JE-Header']
    line_ws = wb['JE-Line']

    ensure_line_taxdate_column(line_ws)

    clear_sheet_data(header_ws, 4, len(HEADER_COLUMNS))
    clear_sheet_data(line_ws, 4, len(LINE_COLUMNS))

    for row_idx, row_data in enumerate(header_rows, start=4):
        for col_idx, col_name in enumerate(HEADER_COLUMNS, start=1):
            cell = header_ws.cell(row_idx, col_idx, row_data.get(col_name))
            if row_data.get('JdtNum') in highlight_header_keys:
                cell.fill = HIGHLIGHT_FILL

    for row_idx, row_data in enumerate(line_rows, start=4):
        line_key = (row_data.get('ParentKey'), row_data.get('LineNum'))
        for col_idx, col_name in enumerate(LINE_COLUMNS, start=1):
            cell = line_ws.cell(row_idx, col_idx, row_data.get(col_name))
            if line_key in highlight_line_keys:
                cell.fill = HIGHLIGHT_FILL

    wb.save(output_path)



def format_txt_value(value: Any, column_name: str) -> str:
    if value is None:
        return ''
    if isinstance(value, datetime):
        if column_name == 'U_Invdate':
            return value.strftime('%d/%m/%Y')
        return value.strftime('%Y%m%d')
    if isinstance(value, date):
        if column_name == 'U_Invdate':
            return value.strftime('%d/%m/%Y')
        return value.strftime('%Y%m%d')
    if isinstance(value, bool):
        return 'Y' if value else 'N'
    if isinstance(value, float):
        if column_name in {'Debit', 'Credit', 'FCDebit', 'FCCredit', 'BaseSum'}:
            return f'{value:.2f}'
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        if column_name in {'Debit', 'Credit', 'FCDebit', 'FCCredit', 'BaseSum'}:
            return f'{value:.2f}'
        return str(value)
    text = str(value)
    return text



def export_sheet_to_txt(workbook_path: Path, sheet_name: str, txt_path: Path, max_col: int) -> None:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    ws = wb[sheet_name]
    lines: list[str] = []
    for row_idx in range(2, ws.max_row + 1):
        values = []
        has_content = False
        for col_idx in range(1, max_col + 1):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value not in (None, ''):
                has_content = True
            column_name = clean_text(ws.cell(2, col_idx).value) or f'COL_{col_idx}'
            values.append(format_txt_value(cell_value, column_name))
        if has_content:
            lines.append('\t'.join(values))
    with txt_path.open('w', encoding='utf-16', newline='') as f:
        f.write('\r\n'.join(lines) + '\r\n')
    wb.close()



def export_txt_outputs(output_path: Path) -> tuple[Path, Path]:
    header_txt_path = output_path.with_name('Header.txt')
    line_txt_path = output_path.with_name('Line.txt')
    export_sheet_to_txt(output_path, 'JE-Header', header_txt_path, len(HEADER_COLUMNS))
    export_sheet_to_txt(output_path, 'JE-Line', line_txt_path, len(LINE_COLUMNS))
    return header_txt_path, line_txt_path



def build_summary(output_path: Path, summary_path: Path, header_txt_path: Path, line_txt_path: Path, template_path: Path, billing_month: str, posting_date: date, gfb_rows: list[dict[str, Any]], artifacts: BuildArtifacts) -> ConversionResult:
    total_amount = round(sum(to_num(r.get('AMOUNT')) for r in gfb_rows))
    result = ConversionResult(
        output_path=str(output_path),
        summary_path=str(summary_path),
        header_txt_path=str(header_txt_path),
        line_txt_path=str(line_txt_path),
        template_path=str(template_path),
        billing_month=billing_month,
        posting_date=posting_date.isoformat(),
        gfb_rows=len(gfb_rows),
        header_rows=len(artifacts.header_rows),
        line_rows=len(artifacts.line_rows),
        department_count=len(artifacts.department_summary),
        total_amount=total_amount,
        total_debit=artifacts.total_debit,
        total_credit=artifacts.total_credit,
        warnings=artifacts.warnings,
        department_summary=artifacts.department_summary,
    )
    summary_path.write_text(json.dumps(result.__dict__, ensure_ascii=False, indent=2), encoding='utf-8')
    return result



def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Convert Grab billing report to SAP JE template workbook')
    parser.add_argument('input_path', help='Path to GFB report file or month folder')
    parser.add_argument('--output', help='Output xlsx path')
    parser.add_argument('--template', help='Template xlsx path')
    parser.add_argument('--posting-date', help='Override posting date dd/mm/yyyy')
    return parser.parse_args()



def main() -> int:
    args = parse_args()
    input_path = Path(args.input_path)
    gfb_path = resolve_input_path(input_path)
    template_path = resolve_template_path(args.template)
    gfb_rows = read_gfb_rows(gfb_path)
    year, month = detect_billing_month(gfb_rows)
    billing_month = f'{month:02d}.{year}'
    posting_date = datetime.strptime(args.posting_date, '%d/%m/%Y').date() if args.posting_date else get_last_day_of_next_month(year, month)
    output_path = resolve_output_path(gfb_path.parent, billing_month, args.output)
    summary_path = output_path.with_name(output_path.stem + ' - Summary.json')

    defaults = load_template_defaults(template_path)
    artifacts = build_rows(gfb_rows, posting_date, billing_month, defaults)
    write_output(template_path, output_path, artifacts.header_rows, artifacts.line_rows, artifacts.highlight_header_keys, artifacts.highlight_line_keys)
    header_txt_path, line_txt_path = export_txt_outputs(output_path)
    result = build_summary(output_path, summary_path, header_txt_path, line_txt_path, template_path, billing_month, posting_date, gfb_rows, artifacts)

    print(f'File GFB           : {gfb_path}')
    print(f'Template           : {template_path}')
    print(f'Posting Date       : {yyyymmdd_int(posting_date)}')
    print(f'Billing month      : {billing_month}')
    print(f'GFB rows           : {artifacts.processed_rows}')
    print(f'JE Header rows     : {len(artifacts.header_rows)}')
    print(f'JE Line rows       : {len(artifacts.line_rows)}')
    print(f'Departments        : {len(artifacts.department_summary)}')
    print(f'Total GFB amount   : {result.total_amount:,.0f}')
    print(f'Total Debit        : {result.total_debit:,.0f}')
    print(f'Total Credit       : {result.total_credit:,.0f}')
    print(f'Output             : {output_path}')
    print(f'Header TXT         : {header_txt_path}')
    print(f'Line TXT           : {line_txt_path}')
    print(f'Summary            : {summary_path}')
    if artifacts.warnings:
        print('Warnings:')
        for w in artifacts.warnings[:20]:
            print(f'- {w}')
        if len(artifacts.warnings) > 20:
            print(f'- ... còn {len(artifacts.warnings) - 20} warning khác')
    return 0 if result.total_amount == result.total_debit == result.total_credit else 1


if __name__ == '__main__':
    raise SystemExit(main())
