"""converter.py - Chuyển đổi hoá đơn điện (Excel nguồn) -> Excel theo mẫu import SAP.

Usage:
  python converter.py <input.xlsx> [--month MM] [--year YYYY] [--output out.xlsx]

- Nếu không truyền --month/--year, script sẽ tự suy từ cột `NGÀY PHÁT HÀNH`.
- Nếu dữ liệu chứa nhiều tháng/năm khác nhau và không truyền --month/--year, script sẽ dừng
  và in ra danh sách các kỳ để user chọn.

Yêu cầu cột input (đúng tên):
  số HD, Ký hiệu, TONG_NO, NGÀY PHÁT HÀNH

Ghi chú môi trường:
- Không phụ thuộc pandas. Dùng openpyxl.
"""

from __future__ import annotations

import argparse
import calendar
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
import sys
from typing import Any

from openpyxl import load_workbook, Workbook

# Fix console encoding on Windows (avoid UnicodeEncodeError on cp1252 consoles)
try:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# =========================
# CONFIG CỐ ĐỊNH
# =========================

CONFIG = {
    "TK_CHI_PHI": {
        "code": 62721001,
        "name": "Xây dựng cơ bản",
        "distr_rule": "12090310;M999994;M02;PMO;M0100000",
    },
    "TK_THUE": {
        "code": 13331001,
        "name": "Thuế GTGT được khấu trừ của Dự án",
        "tax_group": "PVN5",
    },
    "PROJECT": "M02",
    "BRANCH": "LEGACY",
    "MA_DOI_TAC": "V00000162",
    "TEN_DOI_TAC": "CHI NHÁNH TỔNG CÔNG TY ĐIỆN LỰC TPHCM TNHH-CÔNG TY ĐIỆN LỰC SÀI GÒN",
    "MST": "0300951119-001",
    "TINH_TRANG_KE_KHAI": "Kê khai",
    "VAT_RATE": Decimal("0.08"),
}

SAP_COLUMNS = [
    "G/L Acct/BP Code",
    "G/L Acct/BP Name",
    "Control Acct",
    "Debit",
    "Credit",
    "Distr. Rule",
    "Primary Form Item",
    "CFWId",
    "Bank Account",
    "Remarks",
    "Offset Account",
    "Ref. 2",
    "Due Date",
    "Posting Date",
    "Document Date",
    "Project/Khế ước",
    "Tax Group",
    "Federal Tax ID",
    "Tax Amount",
    "Gross Value",
    "Base Amount",
    "Branch",
    "Số HĐ",
    "Seri HĐ",
    "InvType",
    "Tình trạng kê khai",
    "Diễn giải HĐKM",
    "Nhãn tính C.Nợ",
    "Mẫu số HĐ",
    "AdjTran",
    "Mã đối tác",
    "Tên đối tác",
    "Địa chỉ",
    "MST",
    "Diễn giải",
    "RemarksJE",
    "BP Bank Account",
    "Share Holder No",
]

REQUIRED_INPUT_COLUMNS = ["số HD", "Ký hiệu", "TONG_NO", "NGÀY PHÁT HÀNH"]


# =========================
# UTILS
# =========================


def excel_round_half_up(value: Decimal) -> Decimal:
    """Excel ROUND(value, 0) ~ làm tròn half-up tới số nguyên."""
    return value.quantize(Decimal("1"), rounding=ROUND_HALF_UP)


def to_decimal(x: Any) -> Decimal:
    if x is None:
        return Decimal("0")
    if isinstance(x, Decimal):
        return x
    # openpyxl có thể đưa int/float
    try:
        return Decimal(str(x))
    except Exception:
        return Decimal("0")


def parse_date(value: Any) -> datetime:
    """Parse ngày phát hành. Ưu tiên kiểu datetime/date từ Excel; fallback parse string dd/mm/yyyy."""
    if value is None:
        raise ValueError("NGÀY PHÁT HÀNH trống")

    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)

    s = str(value).strip()
    if not s:
        raise ValueError("NGÀY PHÁT HÀNH trống")

    # thử các format phổ biến
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass

    # có thể có kèm time
    for fmt in ("%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            pass

    raise ValueError(f"Không parse được NGÀY PHÁT HÀNH: {value!r}")


def tao_dien_giai(thang: int, nam: int) -> str:
    ngay_cuoi = calendar.monthrange(nam, thang)[1]
    return (
        f"Điện tiêu thụ tháng {thang} năm {nam} "
        f"từ ngày 01/{thang:02d}/{nam} đến ngày {ngay_cuoi:02d}/{thang:02d}/{nam}"
    )


def tinh_tien(tong_no: Any) -> tuple[int, int]:
    tong_no_dec = to_decimal(tong_no)
    divisor = Decimal("1") + CONFIG["VAT_RATE"]  # 1.08

    chi_phi_dec = excel_round_half_up(tong_no_dec / divisor)
    thue_dec = tong_no_dec - chi_phi_dec

    return int(chi_phi_dec), int(thue_dec)


@dataclass
class InvoiceRow:
    so_hd: int
    ky_hieu: str
    tong_no: int
    ngay_phat_hanh: datetime


@dataclass
class Period:
    month: int
    year: int


def read_input_rows(input_path: Path) -> tuple[list[InvoiceRow], list[str]]:
    wb = load_workbook(filename=str(input_path), data_only=True)
    ws = wb.active

    header = []
    for cell in ws[1]:
        header.append(str(cell.value).strip() if cell.value is not None else "")

    col_index = {name: idx for idx, name in enumerate(header)}
    missing = [c for c in REQUIRED_INPUT_COLUMNS if c not in col_index]
    if missing:
        raise ValueError(
            "File input thiếu các cột bắt buộc: " + ", ".join(missing) + ". "
            "Vui lòng kiểm tra lại file nguồn từ điện lực."
        )

    warnings: list[str] = []
    rows: list[InvoiceRow] = []

    for r in range(2, ws.max_row + 1):
        so_hd_val = ws.cell(row=r, column=col_index["số HD"] + 1).value
        tong_no_val = ws.cell(row=r, column=col_index["TONG_NO"] + 1).value

        # bỏ qua dòng rỗng
        if so_hd_val is None or tong_no_val is None:
            continue

        try:
            so_hd = int(to_decimal(so_hd_val))
        except Exception:
            warnings.append(f"Dòng {r}: số HD không hợp lệ ({so_hd_val!r}) → bỏ qua")
            continue

        try:
            tong_no = int(to_decimal(tong_no_val))
        except Exception:
            warnings.append(f"Dòng {r}: TONG_NO không hợp lệ ({tong_no_val!r}) → bỏ qua")
            continue

        ky_hieu_val = ws.cell(row=r, column=col_index["Ký hiệu"] + 1).value
        ky_hieu = "" if ky_hieu_val is None else str(ky_hieu_val).strip()

        ngay_val = ws.cell(row=r, column=col_index["NGÀY PHÁT HÀNH"] + 1).value
        try:
            ngay = parse_date(ngay_val)
        except Exception as e:
            warnings.append(f"Dòng {r}: NGÀY PHÁT HÀNH không hợp lệ ({ngay_val!r}) → bỏ qua ({e})")
            continue

        rows.append(
            InvoiceRow(
                so_hd=so_hd,
                ky_hieu=ky_hieu,
                tong_no=tong_no,
                ngay_phat_hanh=ngay,
            )
        )

    return rows, warnings


def infer_period(rows: list[InvoiceRow]) -> Period:
    periods = sorted({(r.ngay_phat_hanh.month, r.ngay_phat_hanh.year) for r in rows})
    if not periods:
        raise ValueError("Không có dòng hợp lệ để suy ra kỳ (tháng/năm).")
    if len(periods) == 1:
        m, y = periods[0]
        return Period(month=m, year=y)

    msg = "Dữ liệu có nhiều tháng/năm trong cột NGÀY PHÁT HÀNH. Vui lòng chỉ định --month và --year. Các kỳ tìm thấy:\n"
    msg += "\n".join([f"- {m:02d}/{y}" for m, y in periods])
    raise ValueError(msg)


def tao_dong_chi_phi(inv: InvoiceRow, remarks: str) -> dict[str, Any]:
    chi_phi, _ = tinh_tien(inv.tong_no)
    return {
        "G/L Acct/BP Code": CONFIG["TK_CHI_PHI"]["code"],
        "G/L Acct/BP Name": CONFIG["TK_CHI_PHI"]["name"],
        "Control Acct": CONFIG["TK_CHI_PHI"]["code"],
        "Debit": chi_phi,
        "Distr. Rule": CONFIG["TK_CHI_PHI"]["distr_rule"],
        "Remarks": remarks,
        "Due Date": inv.ngay_phat_hanh,
        "Posting Date": inv.ngay_phat_hanh,
        "Document Date": inv.ngay_phat_hanh,
        "Project/Khế ước": CONFIG["PROJECT"],
        "Branch": CONFIG["BRANCH"],
        "Số HĐ": inv.so_hd,
        "Seri HĐ": inv.ky_hieu,
        "Tình trạng kê khai": CONFIG["TINH_TRANG_KE_KHAI"],
        "Mã đối tác": CONFIG["MA_DOI_TAC"],
        "Tên đối tác": CONFIG["TEN_DOI_TAC"],
        "MST": CONFIG["MST"],
        "RemarksJE": remarks,
    }


def tao_dong_thue(inv: InvoiceRow, remarks: str) -> dict[str, Any]:
    _, thue = tinh_tien(inv.tong_no)
    return {
        "G/L Acct/BP Code": CONFIG["TK_THUE"]["code"],
        "G/L Acct/BP Name": CONFIG["TK_THUE"]["name"],
        "Control Acct": CONFIG["TK_THUE"]["code"],
        "Debit": thue,
        "Tax Group": CONFIG["TK_THUE"]["tax_group"],
        "Base Amount": inv.tong_no,
        "Remarks": remarks,
        "Due Date": inv.ngay_phat_hanh,
        "Posting Date": inv.ngay_phat_hanh,
        "Document Date": inv.ngay_phat_hanh,
        "Project/Khế ước": CONFIG["PROJECT"],
        "Branch": CONFIG["BRANCH"],
        "Số HĐ": inv.so_hd,
        "Seri HĐ": inv.ky_hieu,
        "Tình trạng kê khai": CONFIG["TINH_TRANG_KE_KHAI"],
        "Mã đối tác": CONFIG["MA_DOI_TAC"],
        "Tên đối tác": CONFIG["TEN_DOI_TAC"],
        "MST": CONFIG["MST"],
        "RemarksJE": remarks,
    }


def build_output_rows(invoices: list[InvoiceRow], month: int, year: int) -> list[dict[str, Any]]:
    remarks = tao_dien_giai(month, year)
    out: list[dict[str, Any]] = []
    for inv in invoices:
        out.append(tao_dong_chi_phi(inv, remarks))
        out.append(tao_dong_thue(inv, remarks))

    out.sort(key=lambda r: (r.get("Số HĐ") or 0, 0 if r.get("G/L Acct/BP Code") == CONFIG["TK_CHI_PHI"]["code"] else 1))
    return out


def write_output_xlsx(output_path: Path, output_rows: list[dict[str, Any]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SAP_IMPORT"

    # header
    for c, name in enumerate(SAP_COLUMNS, start=1):
        ws.cell(row=1, column=c, value=name)

    # rows
    for r, row_dict in enumerate(output_rows, start=2):
        for c, col_name in enumerate(SAP_COLUMNS, start=1):
            ws.cell(row=r, column=c, value=row_dict.get(col_name))

    wb.save(str(output_path))


def validate_output(invoices: list[InvoiceRow], output_rows: list[dict[str, Any]]) -> tuple[bool, list[str]]:
    errors: list[str] = []

    if len(output_rows) != len(invoices) * 2:
        errors.append(
            f"Số dòng output không đúng: {len(output_rows)} (expected: {len(invoices) * 2})"
        )

    # group by invoice
    by_hd: dict[int, list[dict[str, Any]]] = {}
    for r in output_rows:
        so_hd = int(r.get("Số HĐ") or 0)
        by_hd.setdefault(so_hd, []).append(r)

    for inv in invoices:
        rows = by_hd.get(inv.so_hd, [])
        if len(rows) != 2:
            errors.append(f"HĐ {inv.so_hd}: Không đủ 2 dòng (thực tế {len(rows)})")
            continue

        total_debit = sum(int(to_decimal(x.get("Debit"))) for x in rows)
        if total_debit != inv.tong_no:
            errors.append(f"HĐ {inv.so_hd}: Tổng Debit ({total_debit}) != TONG_NO ({inv.tong_no})")

    return (len(errors) == 0), errors


def build_default_output_name(month: int, year: int) -> str:
    return f"HOA_DON_DIEN_SAP_T{month:02d}{str(year)[-2:]}.xlsx"


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", help="Đường dẫn file Excel nguồn")
    parser.add_argument("--month", type=int, default=None, help="Tháng (1-12)")
    parser.add_argument("--year", type=int, default=None, help="Năm (YYYY)")
    parser.add_argument("--output", default=None, help="Đường dẫn file Excel output")
    parser.add_argument(
        "--no-validate",
        action="store_true",
        help="Bỏ qua bước kiểm tra tổng Debit theo từng hoá đơn",
    )

    args = parser.parse_args(argv)

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ Không tìm thấy file input: {input_path}")
        return 1

    try:
        invoices, warnings = read_input_rows(input_path)
        if not invoices:
            print("❌ Không có dòng hợp lệ sau khi đọc file (thiếu dữ liệu hoặc sai định dạng).")
            return 2

        if args.month is None or args.year is None:
            period = infer_period(invoices)
            month, year = period.month, period.year
        else:
            month, year = int(args.month), int(args.year)

        output_rows = build_output_rows(invoices, month, year)

    except Exception as e:
        print("❌ Lỗi khi chuyển đổi:")
        print(str(e))
        return 2

    out_path = Path(args.output) if args.output else Path(build_default_output_name(month, year))
    out_path.parent.mkdir(parents=True, exist_ok=True)

    write_output_xlsx(out_path, output_rows)

    ok = True
    errors: list[str] = []
    if not args.no_validate:
        ok, errors = validate_output(invoices, output_rows)

    print(f"✅ Đã xuất file: {out_path}")
    print(f"   - Kỳ: {month:02d}/{year}")
    print(f"   - Số hóa đơn gốc (hợp lệ): {len(invoices)}")
    print(f"   - Số dòng SAP: {len(output_rows)}")

    if warnings:
        print(f"⚠️ Cảnh báo: bỏ qua {len(warnings)} dòng do dữ liệu không hợp lệ (hiển thị tối đa 10)")
        for w in warnings[:10]:
            print(f"   - {w}")

    if not ok:
        print("⚠️ Cảnh báo: kiểm tra phát hiện lỗi:")
        for e in errors[:15]:
            print(f"   - {e}")
        print("   (Đã xuất file nhưng cần kiểm tra lại dữ liệu nguồn / cấu hình.)")
        return 3

    print("✅ Kiểm tra thành công!")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
