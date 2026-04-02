#!/usr/bin/env python3
"""
GFB Billing → SAP Import Converter
Chuyển đổi báo cáo cước Grab For Business sang file import SAP B1

Cách dùng:
  python convert_gfb_to_sap.py <input_gfb.xlsx> <output_sap.xlsx> [options]

Options:
  --expense-gl    Tài khoản chi phí (mặc định: 111111111)
  --vat-gl        Tài khoản VAT đầu vào (mặc định: 22222222)
  --vendor-code   Mã nhà cung cấp SAP (mặc định: GRAB)
  --distr-rule    Distr. Rule (mặc định: A00001)
  --tax-group     Nhóm thuế VAT (mặc định: PVN5)
  --posting-date  Ngày hạch toán dd/mm/yyyy (mặc định: cuối tháng chiếm đa số)
"""

import sys
import argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from collections import Counter
import calendar

# ─── Thông tin nhà cung cấp Grab (cố định) ──────────────────────────────────
VENDOR_NAME    = "CÔNG TY TNHH GRAB"
VENDOR_MST     = "0312650437"
VENDOR_ADDRESS = "Tòa nhà Mapletree Business Centre, 1060 Nguyễn Văn Linh, Phường Tân Hưng, TP. Hồ Chí Minh"
BRANCH         = "LEGACY"
TINH_TRANG     = "Kê khai"

SAP_COLUMNS = [
    'G/L Acct/BP Code', 'G/L Acct/BP Name', 'Control Acct',
    'Credit', 'Credit (FC)', 'Debit', 'Debit (FC)',
    'Distr. Rule', 'Primary Form Item', 'CFWId', 'Bank Account',
    'Remarks', 'Offset Account', 'Ref. 2',
    'Due Date', 'Posting Date', 'Document Date',
    'Project/Khế ước', 'Tax Group', 'Tax Amount', 'Base Amount',
    'Federal Tax ID', 'Gross Value', 'Branch',
    'Số HĐ', 'Seri HĐ', 'InvType', 'Tình trạng kê khai',
    'Diễn giải HĐKM', 'Nhãn tính C.Nợ', 'Mẫu số HĐ', 'AdjTran',
    'Mã đối tác', 'Tên đối tác', 'Địa chỉ', 'MST',
    'Diễn giải', 'RemarksJE', 'BP Bank Account', 'Share Holder No'
]


def get_last_day_of_month(year, month):
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, last_day)


def to_num(v):
    try:
        return float(v) if v else 0
    except (ValueError, TypeError):
        return 0


def read_gfb_data(filepath):
    """Đọc file GFB Billing Report"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is not None for cell in row):
            row_dict = dict(zip(headers, row))
            if to_num(row_dict.get('AMOUNT')) > 0:
                rows.append(row_dict)
    return rows


def detect_billing_month(rows):
    """Xác định tháng kỳ thanh toán dựa theo tháng chiếm đa số trong VAT_INVOICE_DATE"""
    months = []
    for r in rows:
        d = r.get('VAT_INVOICE_DATE')
        if d:
            months.append(str(d)[:7])  # 'YYYY-MM'
    if not months:
        today = date.today()
        return today.year, today.month
    most_common = Counter(months).most_common(1)[0][0]
    year, month = map(int, most_common.split('-'))
    return year, month


def aggregate_by_department(rows):
    """Tổng hợp tài chính theo phòng ban"""
    dept_data = {}
    for row in rows:
        dept = (row.get('GROUP_NAME') or 'Unknown').strip()
        if dept not in dept_data:
            dept_data[dept] = {
                'pre_vat': 0, 'vat': 0, 'total': 0,
                'trips': 0, 'serials': set()
            }
        d = dept_data[dept]
        d['pre_vat'] += to_num(row.get('PRE_VAT_DELIVERY_FEE')) + to_num(row.get('PRE_VAT_SERVICE_FEE'))
        d['vat']     += to_num(row.get('VAT_VALUE_DELIVERY_FEE')) + to_num(row.get('VAT_VALUE_SERVICE_FEE'))
        d['total']   += to_num(row.get('AMOUNT'))
        d['trips']   += 1
        if row.get('VAT_INVOICE_SERIAL'):
            d['serials'].add(row['VAT_INVOICE_SERIAL'])
    return dept_data


def build_sap_rows(dept_data, args, posting_date):
    """Tạo dữ liệu SAP: 2 dòng/phòng ban (chi phí + VAT)"""
    sap_rows = []
    month_str = posting_date.strftime('%m/%Y')

    for dept, d in sorted(dept_data.items()):
        pre_vat = round(d['pre_vat'])
        vat_amt = round(d['vat'])
        serials = ', '.join(sorted(d['serials']))
        dien_giai = f"Chi phí đi lại Grab tháng {month_str} - {dept}"

        # Dòng 1: Chi phí (Debit TK chi phí)
        row_exp = {col: None for col in SAP_COLUMNS}
        row_exp.update({
            'G/L Acct/BP Code'  : args.expense_gl,
            'G/L Acct/BP Name'  : f"Chi phí đi lại - {dept}",
            'Control Acct'      : args.expense_gl,
            'Debit'             : pre_vat,
            'Distr. Rule'       : args.distr_rule,
            'Remarks'           : dien_giai,
            'Due Date'          : posting_date,
            'Posting Date'      : posting_date,
            'Document Date'     : posting_date,
            'Branch'            : BRANCH,
            'Seri HĐ'           : serials,
            'Tình trạng kê khai': TINH_TRANG,
            'Mã đối tác'        : args.vendor_code,
            'Tên đối tác'       : VENDOR_NAME,
            'Địa chỉ'           : VENDOR_ADDRESS,
            'MST'               : VENDOR_MST,
            'Diễn giải'         : dien_giai,
        })
        sap_rows.append(row_exp)

        # Dòng 2: Thuế GTGT (Debit TK VAT đầu vào)
        row_vat = {col: None for col in SAP_COLUMNS}
        row_vat.update({
            'G/L Acct/BP Code'  : args.vat_gl,
            'G/L Acct/BP Name'  : "Thuế GTGT đầu vào",
            'Control Acct'      : args.vat_gl,
            'Debit'             : vat_amt,
            'Remarks'           : dien_giai,
            'Due Date'          : posting_date,
            'Posting Date'      : posting_date,
            'Document Date'     : posting_date,
            'Tax Group'         : args.tax_group,
            'Tax Amount'        : vat_amt,
            'Base Amount'       : pre_vat,
            'Branch'            : BRANCH,
            'Seri HĐ'           : serials,
            'Tình trạng kê khai': TINH_TRANG,
            'Mã đối tác'        : args.vendor_code,
            'Tên đối tác'       : VENDOR_NAME,
            'Địa chỉ'           : VENDOR_ADDRESS,
            'MST'               : VENDOR_MST,
            'Diễn giải'         : dien_giai,
        })
        sap_rows.append(row_vat)

    return sap_rows


def write_sap_excel(sap_rows, output_path):
    """Ghi file SAP Import Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "HOA DON"

    # Header
    h_fill = PatternFill("solid", fgColor="4472C4")
    h_font = Font(bold=True, color="FFFFFF", size=10)
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    for ci, col in enumerate(SAP_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=col)
        c.fill = h_fill; c.font = h_font; c.alignment = h_align
    ws.row_dimensions[1].height = 30

    # Data
    exp_fill = PatternFill("solid", fgColor="DEEAF1")
    vat_fill = PatternFill("solid", fgColor="E2EFDA")

    for ri, row_data in enumerate(sap_rows, 2):
        is_vat = row_data.get('Tax Group') is not None
        fill = vat_fill if is_vat else exp_fill
        for ci, col in enumerate(SAP_COLUMNS, 1):
            val = row_data.get(col)
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill
            c.alignment = Alignment(vertical='center')
            if isinstance(val, date) and not isinstance(val, datetime):
                c.number_format = 'DD/MM/YYYY'
            elif col in ['Debit', 'Credit', 'Tax Amount', 'Base Amount']:
                c.number_format = '#,##0'

    # Column widths
    widths = {
        'G/L Acct/BP Code': 15, 'G/L Acct/BP Name': 28, 'Control Acct': 14,
        'Debit': 13, 'Credit': 12, 'Distr. Rule': 12,
        'Remarks': 45, 'Due Date': 12, 'Posting Date': 12, 'Document Date': 12,
        'Tax Group': 10, 'Tax Amount': 12, 'Base Amount': 12,
        'Branch': 8, 'Seri HĐ': 15, 'Tình trạng kê khai': 15,
        'Mã đối tác': 12, 'Tên đối tác': 35, 'MST': 15, 'Diễn giải': 45,
    }
    for ci, col in enumerate(SAP_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col, 10)

    ws.freeze_panes = 'A2'
    wb.save(output_path)


def print_summary(dept_data, posting_date, output_path):
    """In bảng tóm tắt kết quả"""
    print(f"\n{'─'*70}")
    print(f"{'PHÒNG BAN':<12} {'CHUYẾN':>6}  {'CHI PHÍ (VND)':>15}  {'VAT':>12}  {'TỔNG':>15}")
    print(f"{'─'*70}")
    grand_exp = grand_vat = grand_tot = 0
    for dept, d in sorted(dept_data.items()):
        pre = round(d['pre_vat']); vat = round(d['vat']); tot = round(d['total'])
        grand_exp += pre; grand_vat += vat; grand_tot += tot
        print(f"{dept:<12} {d['trips']:>6}  {pre:>15,.0f}  {vat:>12,.0f}  {tot:>15,.0f}")
    print(f"{'─'*70}")
    print(f"{'TỔNG CỘNG':<12} {'':>6}  {grand_exp:>15,.0f}  {grand_vat:>12,.0f}  {grand_tot:>15,.0f}")
    print(f"{'─'*70}")
    print(f"\n📅 Ngày hạch toán : {posting_date.strftime('%d/%m/%Y')}")
    print(f"📄 Số dòng SAP     : {len(dept_data) * 2} dòng ({len(dept_data)} phòng ban × 2)")
    print(f"💾 File output     : {output_path}")


def main():
    parser = argparse.ArgumentParser(description='GFB Billing → SAP Import')
    parser.add_argument('input')
    parser.add_argument('output')
    parser.add_argument('--expense-gl',   default='111111111')
    parser.add_argument('--vat-gl',       default='22222222')
    parser.add_argument('--vendor-code',  default='GRAB')
    parser.add_argument('--distr-rule',   default='A00001')
    parser.add_argument('--tax-group',    default='PVN5')
    parser.add_argument('--posting-date', default=None,
                        help='dd/mm/yyyy – mặc định: cuối tháng chiếm đa số')
    args = parser.parse_args()

    print(f"📂 Đọc file: {args.input}")
    rows = read_gfb_data(args.input)
    print(f"   → {len(rows)} chuyến đi hợp lệ")

    if args.posting_date:
        posting_date = datetime.strptime(args.posting_date, '%d/%m/%Y').date()
    else:
        year, month = detect_billing_month(rows)
        posting_date = get_last_day_of_month(year, month)

    dept_data = aggregate_by_department(rows)
    sap_rows  = build_sap_rows(dept_data, args, posting_date)
    write_sap_excel(sap_rows, args.output)
    print_summary(dept_data, posting_date, args.output)
    return 0


if __name__ == '__main__':
    sys.exit(main())
